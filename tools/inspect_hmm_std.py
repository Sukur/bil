#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook(
    '/Users/shukurrzayev/Documents/bil/bil/docs/CJICM Logistics GmbH - HMM Ratesheet Germany April 2026.xlsx',
    data_only=True
)
print('Sheets:', wb.sheetnames)
for sn in wb.sheetnames[:6]:
    sheet = wb[sn]
    print(f'\n=== {sn} (max_row={sheet.max_row}, max_col={sheet.max_column}) ===')
    count = 0
    for row_idx in range(1, sheet.max_row+1):
        row = sheet[row_idx]
        cells = [c.value for c in row[:26]]
        non_empty = [x for x in cells if x is not None]
        if non_empty:
            print(f'  row{row_idx:3d}: {[str(v)[:20] if v else None for v in cells[:8]]}')
            count += 1
        if count >= 20:
            print('  ...')
            break

