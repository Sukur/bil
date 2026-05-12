#!/usr/bin/env python3
"""Inspect Excel structure of failing parser files."""
import openpyxl

FILES = {
    "OOG": "/Users/shukurrzayev/Documents/bil/bil/docs/COSCO SHIPPING LINES GERMANY - FAK RATESHEET OOG - 01.04. - 30.04.2026 - V2.xlsx",
    "REEFER": "/Users/shukurrzayev/Documents/bil/bil/docs/COSCO SHIPPING LINES GERMANY - FAK RATESHEET REEFER - 01.04. - 30.04.2026 - V1.xlsx",
    "HMM_SPECIAL": "/Users/shukurrzayev/Documents/bil/bil/docs/CJ ICM Logistics GmbH - HMM GER Special Ratesheet APR 2026.xlsx",
    "OOCL": "/Users/shukurrzayev/Documents/bil/bil/docs/REVISED OOCL Export FAK Fareast_Mideast_India_Australia_Intra Europe_ Africa 01.12.2025 - 31.01.2026.xlsx",
}

for label, path in FILES.items():
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"\n{'='*70}")
    print(f"FILE: {label}")
    print(f"SHEETS: {wb.sheetnames}")
    # Show first and second sheet structure
    for sheet_name in wb.sheetnames[:3]:
        sheet = wb[sheet_name]
        print(f"\n  --- Sheet: '{sheet_name}' (max_row={sheet.max_row}, max_col={sheet.max_column}) ---")
        count = 0
        for row_idx in range(1, 35):
            row = sheet[row_idx]
            cells = [c.value for c in row[:14]]
            non_empty = [x for x in cells if x is not None]
            if non_empty:
                # Show first 12 columns
                print(f"    row{row_idx:2d}: {[str(v)[:30] if v else None for v in cells[:14]]}")
                count += 1
            if count >= 20:
                break

