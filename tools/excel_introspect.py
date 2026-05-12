#!/usr/bin/env python3
"""
Excel ratesheet introspection tool.

Reads every .xlsx in docs/ and writes a JSON describing each sheet:
  - sheet name
  - merged regions
  - first ~30 rows (cells with row, col, value)
  - detected header row (first row with >= 3 non-empty string cells)

Output: docs/_introspection/<filename>.json

Usage:
  pip install openpyxl
  python tools/excel_introspect.py
"""
from __future__ import annotations

import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("openpyxl not installed. Run: pip install openpyxl")


ROOT = Path(__file__).resolve().parent.parent
DOCS = ROOT / "docs"
OUT = DOCS / "_introspection"
MAX_ROWS = 40
MAX_COLS = 30


def cell_to_str(v):
    if v is None:
        return None
    if isinstance(v, (int, float, bool, str)):
        return v
    return str(v)


def detect_header_row(rows):
    for r in rows:
        non_empty_strings = [c for c in r["cells"] if isinstance(c["value"], str) and c["value"].strip()]
        if len(non_empty_strings) >= 3:
            return r["row"]
    return None


def introspect_file(path: Path) -> dict:
    wb = load_workbook(path, data_only=True, read_only=False)
    sheets = []
    for ws in wb.worksheets:
        rows = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=MAX_ROWS, max_col=MAX_COLS), start=1):
            cells = []
            for c_idx, cell in enumerate(row, start=1):
                v = cell_to_str(cell.value)
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue
                cells.append({"row": r_idx, "col": c_idx, "value": v})
            if cells:
                rows.append({"row": r_idx, "cells": cells})
        merged = [str(mr) for mr in ws.merged_cells.ranges]
        sheets.append({
            "name": ws.title,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged_regions": merged[:50],
            "rows_preview": rows,
            "header_row_guess": detect_header_row(rows),
        })
    return {"file": path.name, "sheets": sheets}


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    files = sorted(DOCS.glob("*.xlsx"))
    if not files:
        print("No .xlsx files in docs/")
        return
    summary = []
    for f in files:
        print(f"-> {f.name}")
        try:
            data = introspect_file(f)
        except Exception as e:
            print(f"   ERROR: {e}")
            summary.append({"file": f.name, "error": str(e)})
            continue
        out_path = OUT / (f.stem + ".json")
        out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2))
        summary.append({
            "file": f.name,
            "sheets": [{"name": s["name"], "dim": s["dimensions"], "header_row": s["header_row_guess"]} for s in data["sheets"]],
        })
    (OUT / "_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"\nWrote {len(files)} files into {OUT}")


if __name__ == "__main__":
    main()

