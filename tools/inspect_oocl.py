#!/usr/bin/env python3
"""Deep-inspect OOCL Far East sheet and CoscoReefer to understand parser failures."""
import openpyxl, traceback

def inspect(path, sheet_name, max_rows=30):
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        sheet = wb[sheet_name]
    except KeyError:
        print(f"  Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return
    print(f"\n  Sheet '{sheet_name}' max_row={sheet.max_row} max_col={sheet.max_column}")
    count = 0
    for row_idx in range(1, sheet.max_row + 1):
        row = sheet[row_idx]
        cells = [c.value for c in row[:16]]
        non_empty = [x for x in cells if x is not None]
        if non_empty:
            print(f"    row{row_idx:2d}: {[str(v)[:25] if v else None for v in cells]}")
            count += 1
        if count >= max_rows:
            print(f"    ... (stopped at {max_rows} non-empty rows)")
            break

print("="*70)
print("OOCL Far East+Japan")
inspect(
    "/Users/shukurrzayev/Documents/bil/bil/docs/REVISED OOCL Export FAK Fareast_Mideast_India_Australia_Intra Europe_ Africa 01.12.2025 - 31.01.2026.xlsx",
    "Far East+Japan"
)
print("="*70)
print("OOCL Intra Europe")
inspect(
    "/Users/shukurrzayev/Documents/bil/bil/docs/REVISED OOCL Export FAK Fareast_Mideast_India_Australia_Intra Europe_ Africa 01.12.2025 - 31.01.2026.xlsx",
    "Intra Europe"
)
print("="*70)
print("COSCO Reefer FE")
inspect(
    "/Users/shukurrzayev/Documents/bil/bil/docs/COSCO SHIPPING LINES GERMANY - FAK RATESHEET REEFER - 01.04. - 30.04.2026 - V1.xlsx",
    "Reefer FE"
)

